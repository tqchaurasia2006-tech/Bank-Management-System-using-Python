from openpyxl import load_workbook

def Create_Account():
    
    wb = load_workbook("PROJECT.xlsx")
    sheet = wb.active
    
    name = input("Enter name : ")
    
    contact = input("Enter Contact : ")
    
    while True:
        pin = input("Enter 4-digit PIN : ")
        
        if pin.isdigit() and len(pin) == 4:
            break
        print("Invalid PIN!")
    
    account = 1001
    if sheet.max_row > 1:
        account = sheet.cell(row=sheet.max_row, column=1).value+1
        
    balance = 500
    

    sheet.append([
        account,
        name,
        contact,
        pin,
        balance
    ])
    wb.save("PROJECT.xlsx")
    
    print("\n==========================")
    print("Account Created Successfully")
    print("============================")
    print("Name:",name)
    print("Account No. :",account)
    print("Opening Balance :",balance)
    
def login():

    wb = load_workbook("PROJECT.xlsx")
    sheet = wb.active

    print("\n======LOGIN======")

    account = int(input("Enter Account Number : "))
    pin = input("Enter PIN : ")

    for row in range(2,sheet.max_row + 1):
        acc = sheet.cell(row,1).value
        password = str(sheet.cell(row,4).value)

        if account == acc and pin == password:

            print("\nLogin Successful")

            print("Welcome",sheet.cell(row,2).value)

            User_menu(row)
            
            return

    print("\nInvalid Account Number or PIN")

    return None


def view_balance(row):

    wb = load_workbook("PROJECT.xlsx")
    sheet = wb.active

    name = sheet.cell(row,2).value
    balance = sheet.cell(row,5).value

    print("\n========================")
    print("ACCOUNT DETAILS")
    print("========================")
    print("Name :",name)
    print("Balance : ₹",balance)
    
                          
def User_menu(row):
    while True:

        print("\n======================")
        print("======USER  MENU======")
        print("======================")
        print("1.Deposit Money")
        print("2.Withdrawal Money")
        print("3.View balance")
        print("4.Logout")

        choice = input("Enter Choice : ")
        
        if choice == "1":
            Deposit_Money(row)

        elif choice=="2":
            Withdrawal_Money(row)

        elif choice == "3":
            view_balance(row)

        elif choice == "4":
            print("Logged Out Successfully")
            break

        else:
            print("Invalid Choice")
            

def Deposit_Money(row):
    wb = load_workbook("PROJECT.xlsx")
    sheet = wb.active

    amount = float(input("Enter Amount to Deposit : ₹"))
    if amount <= 0:
        print("Invalid Amount!")
        return
        
    balance = sheet.cell(row,5).value
    balance = balance + amount
    sheet.cell(row,5).value = balance
        
    wb.save("PROJECT.xlsx")
    print("₹",amount,"Deposited Successfully")
    print("New Balance : ₹",balance)
        

def Withdrawal_Money(row):
    wb = load_workbook("PROJECT.xlsx")
    sheet = wb.active
        
    amount = float(input("Enter Amount to Withdraw : ₹"))
    if amount <= 0:
        print("Invalid Amount!")
        return
            
    balance = sheet.cell(row,5).value
    if amount > balance:
        print("\nInsufficient Balance!")
        print("Current Balance : ₹", balance)
        return

    balance = balance - amount

    sheet.cell(row,5).value = balance

    wb.save("PROJECT.xlsx")
    print("\nWithdrawal Successful")
    print("Remaining Balance : ₹", balance)
   
def main_menu():
    while True:
        print("="*40)
        print("=========BANK MANAGEMENT SYSTEM=========")
        print("="*40)
        print("1.Create Account")
        print("2.Login")
        print("3.Exit")

        choice = input("Enter the Choice: ")
        if choice == "1":
            Create_Account()
        elif choice == "2":
            login()
        elif choice == "3":
            print("\nThank You for choosing our Bank...")
            break
        else:
            print("Enter Valid Choice")
main_menu()
